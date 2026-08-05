import openpyxl
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.utils import timezone
from django.db.models import Q, Count, Avg, F
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, FileResponse
from django.template.loader import render_to_string
from django.core.files.base import ContentFile

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from .permissions import IsMD

from xhtml2pdf import pisa

from .models import (
    User, Project, Team, ProjectDocument, ProjectComment, ProjectAuditLog, Notification, Client, Task
)
from .serializers import (
    UserSerializer, ProjectSerializer, TeamSerializer, ProjectDocumentSerializer,
    ProjectCommentSerializer, ProjectAuditLogSerializer, NotificationSerializer, ProjectDetailSerializer
)

# Helper functions
def log_project_action(project, action, user, details=None):
    ProjectAuditLog.objects.create(
        project=project,
        action=action,
        user=user,
        details=details
    )

def notify_user(recipient, title, message):
    Notification.objects.create(
        recipient=recipient,
        title=title,
        message=message
    )

class ProjectAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        role = user.role

        # Permissions filtering
        if role in ['HR', 'MD']:
            projects = Project.objects.all()
        elif role == 'Manager':
            projects = Project.objects.filter(assigned_manager=user)
        elif role == 'TeamLead':
            projects = Project.objects.filter(Q(assigned_teams__lead=user))
        elif role == 'Employee':
            projects = Project.objects.filter(Q(assigned_teams__members=user))
        else:
            projects = Project.objects.none()

        projects = projects.distinct()

        # Archive filtering
        include_archived = request.query_params.get('include_archived', 'false') == 'true'
        if not include_archived:
            projects = projects.filter(is_archived=False)

        # Filters
        status_filter = request.query_params.get('status', '')
        if status_filter:
            projects = projects.filter(status=status_filter)

        priority_filter = request.query_params.get('priority', '')
        if priority_filter:
            projects = projects.filter(priority=priority_filter)

        manager_filter = request.query_params.get('manager_id', '')
        if manager_filter:
            projects = projects.filter(assigned_manager_id=manager_filter)

        client_filter = request.query_params.get('client_name', '')
        if client_filter:
            projects = projects.filter(Q(client_name__icontains=client_filter) | Q(client__name__icontains=client_filter))

        category_filter = request.query_params.get('category', '')
        if category_filter:
            projects = projects.filter(project_category__icontains=category_filter)

        # Search query
        search_query = request.query_params.get('search', '')
        if search_query:
            projects = projects.filter(
                Q(name__icontains=search_query) |
                Q(project_code__icontains=search_query) |
                Q(description__icontains=search_query)
            )

        projects = projects.order_by('-created_at')
        serializer = ProjectSerializer(projects, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        user = request.user
        if user.role not in ['HR', 'MD']:
            return Response({"detail": "Access Denied: Only HR or MD can create projects."}, status=status.HTTP_403_FORBIDDEN)

        name = request.data.get('project_name') or request.data.get('name')
        description = request.data.get('description', '')
        client_name = request.data.get('client_name', '')
        client_contact = request.data.get('client_contact', '')
        project_category = request.data.get('project_category', '')
        priority = request.data.get('priority', 'Medium')
        startdate = request.data.get('start_date') or request.data.get('startdate')
        deadline = request.data.get('end_date') or request.data.get('deadline')
        estimated_budget = request.data.get('estimated_budget')
        technology_stack = request.data.get('technology_stack', '')
        project_color = request.data.get('project_color', '#3b82f6')
        assigned_manager_id = request.data.get('assigned_manager')

        if not name:
            return Response({"detail": "Project name is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Resolve manager if provided
        assigned_manager = None
        if assigned_manager_id:
            try:
                assigned_manager = User.objects.get(id=assigned_manager_id, role='Manager')
            except User.DoesNotExist:
                return Response({"detail": "Selected Project Manager is invalid or not in Manager role."}, status=status.HTTP_400_BAD_REQUEST)

        # Handle Logo Upload
        project_logo = request.FILES.get('project_logo')

        project = Project.objects.create(
            name=name,
            description=description,
            client_name=client_name,
            client_contact=client_contact,
            project_category=project_category,
            priority=priority,
            startdate=startdate or None,
            deadline=deadline or None,
            estimated_budget=estimated_budget or None,
            technology_stack=technology_stack,
            project_color=project_color,
            assigned_manager=assigned_manager,
            created_by=user,
            project_logo=project_logo
        )

        log_project_action(project, "Project Created", user, f"Project '{name}' was initialized.")

        # Notifications
        notify_user(user, "Project Created", f"Project '{project.name}' has been created successfully.")
        if assigned_manager:
            notify_user(assigned_manager, "Manager Assigned", f"You have been assigned as Project Manager for project '{project.name}'.")
            log_project_action(project, "Manager Assigned", user, f"Manager {assigned_manager.username} was assigned.")

        # Notify all HR and MD users
        for admin_user in User.objects.filter(role__in=['HR', 'MD']).exclude(id=user.id):
            notify_user(admin_user, "Project Created", f"A new project '{project.name}' has been created.")

        serializer = ProjectSerializer(project, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ProjectDetailAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get_project(self, pk, user):
        project = get_object_or_404(Project, pk=pk)
        role = user.role

        # Check permissions
        if role in ['HR', 'MD']:
            return project
        if role == 'Manager' and project.assigned_manager == user:
            return project
        if role == 'TeamLead' and project.assigned_teams.filter(lead=user).exists():
            return project
        if role == 'Employee' and project.assigned_teams.filter(members=user).exists():
            return project

        raise PermissionError("Access Denied: You are not authorized to view this project.")

    def get(self, request, pk):
        try:
            project = self.get_project(pk, request.user)
            serializer = ProjectDetailSerializer(project, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        except PermissionError as e:
            return Response({"detail": str(e)}, status=status.HTTP_403_FORBIDDEN)

    def put(self, request, pk):
        user = request.user
        role = user.role

        try:
            project = get_object_or_404(Project, pk=pk)
            # Permission check: MD/HR or assigned Project Manager
            if role not in ['HR', 'MD'] and project.assigned_manager != user:
                return Response({"detail": "Access Denied: Only MD, HR, or the assigned Manager can update project details."}, status=status.HTTP_403_FORBIDDEN)

            name = request.data.get('project_name') or request.data.get('name')
            description = request.data.get('description')
            startdate = request.data.get('start_date') or request.data.get('startdate')
            deadline = request.data.get('end_date') or request.data.get('deadline')
            client_name = request.data.get('client_name')
            client_contact = request.data.get('client_contact')
            project_category = request.data.get('project_category')
            priority = request.data.get('priority')
            estimated_budget = request.data.get('estimated_budget')
            technology_stack = request.data.get('technology_stack')
            project_color = request.data.get('project_color')
            status_val = request.data.get('status')
            assigned_manager_id = request.data.get('assigned_manager')

            # Fields MD/HR can modify but Project Manager cannot
            if role in ['HR', 'MD']:
                if name:
                    project.name = name
                if estimated_budget is not None:
                    project.estimated_budget = estimated_budget or None
                if assigned_manager_id is not None:
                    if assigned_manager_id == "" or assigned_manager_id == "null" or assigned_manager_id is None:
                        if project.assigned_manager:
                            log_project_action(project, "Manager Removed", user, f"Manager {project.assigned_manager.username} was unassigned.")
                            notify_user(project.assigned_manager, "Manager Unassigned", f"You have been unassigned from project '{project.name}'.")
                            project.assigned_manager = None
                    else:
                        try:
                            new_manager = User.objects.get(id=assigned_manager_id, role='Manager')
                            if project.assigned_manager != new_manager:
                                old_mgr = project.assigned_manager
                                project.assigned_manager = new_manager
                                log_project_action(project, "Manager Changed", user, f"Manager changed to {new_manager.username}.")
                                notify_user(new_manager, "Manager Assigned", f"You have been assigned as Manager for project '{project.name}'.")
                                if old_mgr:
                                    notify_user(old_mgr, "Manager Unassigned", f"You have been unassigned from project '{project.name}'.")
                        except User.DoesNotExist:
                            return Response({"detail": "Invalid manager selection."}, status=status.HTTP_400_BAD_REQUEST)

            # Common fields both PM and HR/MD can modify
            if description is not None:
                project.description = description
            if startdate is not None:
                project.startdate = startdate or None
            if deadline is not None:
                project.deadline = deadline or None
            if client_name is not None:
                project.client_name = client_name
            if client_contact is not None:
                project.client_contact = client_contact
            if project_category is not None:
                project.project_category = project_category
            if priority is not None:
                project.priority = priority
            if technology_stack is not None:
                project.technology_stack = technology_stack
            if project_color is not None:
                project.project_color = project_color

            # Status change
            if status_val is not None:
                old_status = project.status
                if old_status != status_val:
                    project.status = status_val
                    log_project_action(project, "Project Updated", user, f"Status changed from '{old_status}' to '{status_val}'.")
                    
                    # Notify manager, leads, members of status update
                    recipients = set()
                    if project.assigned_manager:
                        recipients.add(project.assigned_manager)
                    for team in project.assigned_teams.all():
                        if team.lead:
                            recipients.add(team.lead)
                        for member in team.members.all():
                            recipients.add(member)
                    
                    for r in recipients:
                        notify_user(r, "Project Updated", f"Project '{project.name}' status has changed to '{status_val}'.")

                    if status_val == 'Completed':
                        log_project_action(project, "Project Completed", user, f"Project '{project.name}' marked completed.")

            # Logo Upload
            if request.FILES.get('project_logo'):
                project.project_logo = request.FILES['project_logo']

            project.save()
            log_project_action(project, "Project Updated", user, "Project details updated.")
            serializer = ProjectSerializer(project, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)

        except PermissionError as e:
            return Response({"detail": str(e)}, status=status.HTTP_403_FORBIDDEN)

    def delete(self, request, pk):
        user = request.user
        if user.role not in ['HR', 'MD']:
            return Response({"detail": "Access Denied: Only HR or MD can delete projects."}, status=status.HTTP_403_FORBIDDEN)

        project = get_object_or_404(Project, pk=pk)
        project_name = project.name
        project.delete()

        # Create global log
        for admin_user in User.objects.filter(role__in=['HR', 'MD']).exclude(id=user.id):
            notify_user(admin_user, "Project Deleted", f"Project '{project_name}' has been deleted by {user.username}.")

        return Response({"detail": "Project deleted successfully."}, status=status.HTTP_204_NO_CONTENT)


class ProjectArchiveAPIView(APIView):
    permission_classes = [IsAuthenticated, IsMD]

    def post(self, request, pk):
        user = request.user
        if user.role != 'MD':
            return Response({"detail": "Access Denied: Only MD can archive projects."}, status=status.HTTP_403_FORBIDDEN)

        project = get_object_or_404(Project, pk=pk)
        archive_val = request.data.get('archive', True)

        project.is_archived = archive_val
        if archive_val:
            project.status = 'Archived'
            log_project_action(project, "Project Completed", user, "Project archived.")
            notify_msg = f"Project '{project.name}' has been archived."
        else:
            project.status = 'Active'
            log_project_action(project, "Project Updated", user, "Project unarchived.")
            notify_msg = f"Project '{project.name}' has been unarchived."

        project.save()

        # Notify manager and admin users
        recipients = set(User.objects.filter(role__in=['HR', 'MD']))
        if project.assigned_manager:
            recipients.add(project.assigned_manager)

        for r in recipients:
            notify_user(r, "Project Archived Toggle", notify_msg)

        return Response({"detail": "Project archive status updated successfully.", "is_archived": project.is_archived}, status=status.HTTP_200_OK)


class ProjectTransferAPIView(APIView):
    permission_classes = [IsAuthenticated, IsMD]

    def post(self, request, pk):
        user = request.user
        if user.role != 'MD':
            return Response({"detail": "Access Denied: Only MD can transfer projects."}, status=status.HTTP_403_FORBIDDEN)

        project = get_object_or_404(Project, pk=pk)
        new_manager_id = request.data.get('manager_id')

        if not new_manager_id:
            return Response({"detail": "New manager_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        new_manager = get_object_or_404(User, id=new_manager_id, role='Manager')
        old_manager = project.assigned_manager

        if old_manager == new_manager:
            return Response({"detail": "Project is already assigned to this manager."}, status=status.HTTP_400_BAD_REQUEST)

        project.assigned_manager = new_manager
        project.save()

        # Logs
        log_project_action(
            project,
            "Manager Changed",
            user,
            f"Transferred from {old_manager.username if old_manager else 'None'} to {new_manager.username}."
        )

        # Notify
        notify_user(new_manager, "Project Transferred", f"Project '{project.name}' has been transferred to you as Manager.")
        if old_manager:
            notify_user(old_manager, "Project Transferred Away", f"Project '{project.name}' has been transferred away to another Manager.")

        return Response({"detail": "Project transferred successfully."}, status=status.HTTP_200_OK)


class TeamListCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        project = get_object_or_404(Project, id=project_id)
        teams = project.assigned_teams.all()
        serializer = TeamSerializer(teams, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, project_id):
        user = request.user
        project = get_object_or_404(Project, id=project_id)

        # Permissions: MD, HR, or the assigned Manager
        if user.role not in ['HR', 'MD'] and project.assigned_manager != user:
            return Response({"detail": "Access Denied: Only MD, HR, or the assigned Project Manager can create teams."}, status=status.HTTP_403_FORBIDDEN)

        name = request.data.get('name')
        lead_id = request.data.get('lead')
        department = request.data.get('department', 'python_dev')
        description = request.data.get('description', '')
        max_size = request.data.get('max_size', 10)

        if not name:
            return Response({"detail": "Team Name is required."}, status=status.HTTP_400_BAD_REQUEST)

        lead = None
        if lead_id:
            try:
                lead = User.objects.get(id=lead_id, role='TeamLead')
            except User.DoesNotExist:
                return Response({"detail": "Selected Team Lead is invalid."}, status=status.HTTP_400_BAD_REQUEST)

        team = Team.objects.create(
            name=name,
            lead=lead,
            department=department,
            description=description,
            max_size=max_size
        )
        project.assigned_teams.add(team)

        log_project_action(project, "Team Created", user, f"Team '{name}' was created inside the project.")
        notify_user(user, "Team Created", f"Team '{name}' has been successfully created under '{project.name}'.")

        if lead:
            notify_user(lead, "Team Leader Assigned", f"You have been assigned as Team Lead for '{name}' under project '{project.name}'.")
            log_project_action(project, "Team Leader Changed", user, f"Team Lead for '{name}' was set to {lead.username}.")

        serializer = TeamSerializer(team)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class TeamDetailAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        user = request.user
        team = get_object_or_404(Team, pk=pk)
        project = team.projects.first()

        # Permissions
        if user.role not in ['HR', 'MD']:
            if project and project.assigned_manager != user:
                return Response({"detail": "Access Denied: Only MD, HR, or the assigned Manager can update team settings."}, status=status.HTTP_403_FORBIDDEN)
            elif not project and user.role != 'Manager':
                return Response({"detail": "Access Denied."}, status=status.HTTP_403_FORBIDDEN)

        name = request.data.get('name')
        lead_id = request.data.get('lead')
        department = request.data.get('department')
        description = request.data.get('description')
        max_size = request.data.get('max_size')
        is_active = request.data.get('is_active')

        if name:
            team.name = name
        if department:
            team.department = department
        if description is not None:
            team.description = description
        if max_size:
            team.max_size = max_size
        if is_active is not None:
            team.is_active = is_active

        if lead_id is not None:
            if lead_id == "" or lead_id == "null" or lead_id is None:
                if team.lead:
                    if project:
                        log_project_action(project, "Team Leader Changed", user, f"Team Lead for '{team.name}' was removed.")
                    notify_user(team.lead, "Team Lead Unassigned", f"You have been unassigned as Team Lead for '{team.name}'.")
                    team.lead = None
            else:
                try:
                    new_lead = User.objects.get(id=lead_id, role='TeamLead')
                    if team.lead != new_lead:
                        old_lead = team.lead
                        team.lead = new_lead
                        if project:
                            log_project_action(project, "Team Leader Changed", user, f"Team Lead for '{team.name}' set to {new_lead.username}.")
                            notify_user(new_lead, "Team Leader Assigned", f"You have been assigned as Team Lead for '{team.name}' under project '{project.name}'.")
                        else:
                            notify_user(new_lead, "Team Leader Assigned", f"You have been assigned as Team Lead for '{team.name}'.")
                        if old_lead:
                            notify_user(old_lead, "Team Lead Unassigned", f"You have been unassigned as Team Lead for '{team.name}'.")
                except User.DoesNotExist:
                    return Response({"detail": "Invalid Team Lead selection."}, status=status.HTTP_400_BAD_REQUEST)

        team.save()
        if project:
            log_project_action(project, "Team Updated", user, f"Team '{team.name}' was updated.")
        serializer = TeamSerializer(team)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        user = request.user
        team = get_object_or_404(Team, pk=pk)
        project = team.projects.first()

        # Permissions
        if user.role not in ['HR', 'MD']:
            if project and project.assigned_manager != user:
                return Response({"detail": "Access Denied: Only MD, HR, or the assigned Manager can delete teams."}, status=status.HTTP_403_FORBIDDEN)
            elif not project and user.role != 'Manager':
                return Response({"detail": "Access Denied."}, status=status.HTTP_403_FORBIDDEN)

        team_name = team.name
        team.delete()
        if project:
            log_project_action(project, "Team Deleted", user, f"Team '{team_name}' was deleted.")
        return Response({"detail": "Team deleted successfully."}, status=status.HTTP_204_NO_CONTENT)


class TeamMemberAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        user = request.user
        team = get_object_or_404(Team, pk=pk)
        project = team.projects.first()

        # Permissions
        if user.role not in ['HR', 'MD'] and project.assigned_manager != user:
            return Response({"detail": "Access Denied: Only MD, HR, or the assigned Manager can assign team members."}, status=status.HTTP_403_FORBIDDEN)

        employee_id = request.data.get('employee_id')
        if not employee_id:
            return Response({"detail": "employee_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = User.objects.get(id=employee_id, role='Employee')
        except User.DoesNotExist:
            return Response({"detail": "Selected user is invalid or not in Employee role."}, status=status.HTTP_400_BAD_REQUEST)

        # Limit checks
        if team.members.count() >= team.max_size:
            return Response({"detail": f"Team limit reached ({team.max_size} members maximum)."}, status=status.HTTP_400_BAD_REQUEST)

        if team.members.filter(id=employee.id).exists():
            return Response({"detail": "Employee is already in this team."}, status=status.HTTP_400_BAD_REQUEST)

        # Duplicate project assignment check:
        # Check if the employee is already assigned to a team belonging to any active project (Pending, Active, Delayed status)
        active_assignment = Team.objects.filter(
            members=employee,
            project__status__in=['Pending', 'Active', 'Delayed']
        ).first()

        if active_assignment:
            return Response({
                "detail": f"Employee is already assigned to active project '{active_assignment.project.name}' under team '{active_assignment.name}'."
            }, status=status.HTTP_400_BAD_REQUEST)

        team.members.add(employee)
        log_project_action(project, "Employee Assigned", user, f"Employee '{employee.username}' assigned to team '{team.name}'.")
        
        # Notify employee
        notify_user(
            employee,
            "Project Assigned",
            f"You have been assigned to Team '{team.name}' under Project '{project.name}' by {user.username}."
        )

        # Update employee's team name field to match
        employee.team_name = team.name
        employee.save()

        serializer = TeamSerializer(team)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        user = request.user
        team = get_object_or_404(Team, pk=pk)
        project = team.projects.first()

        # Permissions
        if user.role not in ['HR', 'MD'] and project.assigned_manager != user:
            return Response({"detail": "Access Denied: Only MD, HR, or the assigned Manager can remove team members."}, status=status.HTTP_403_FORBIDDEN)

        employee_id = request.query_params.get('employee_id') or request.data.get('employee_id')
        if not employee_id:
            return Response({"detail": "employee_id query parameter or request body is required."}, status=status.HTTP_400_BAD_REQUEST)

        employee = get_object_or_404(User, id=employee_id)

        if not team.members.filter(id=employee.id).exists():
            return Response({"detail": "Employee is not in this team."}, status=status.HTTP_400_BAD_REQUEST)

        team.members.remove(employee)
        log_project_action(project, "Employee Removed", user, f"Employee '{employee.username}' removed from team '{team.name}'.")
        
        # Notify
        notify_user(
            employee,
            "Project Removed",
            f"You have been removed from Team '{team.name}' under Project '{project.name}'."
        )

        # Clear employee's team name field
        if employee.team_name == team.name:
            employee.team_name = ""
            employee.save()

        serializer = TeamSerializer(team)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ProjectCommentAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        comments = project.project_comments.all().order_by('-created_at')
        serializer = ProjectCommentSerializer(comments, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        content = request.data.get('content')

        if not content:
            return Response({"detail": "Comment content cannot be empty."}, status=status.HTTP_400_BAD_REQUEST)

        comment = ProjectComment.objects.create(
            project=project,
            author=request.user,
            content=content
        )

        # Notify project manager and creator if not them
        recipients = set()
        if project.assigned_manager and project.assigned_manager != request.user:
            recipients.add(project.assigned_manager)
        if project.created_by and project.created_by != request.user:
            recipients.add(project.created_by)

        for r in recipients:
            notify_user(r, "New Project Comment", f"{request.user.username} commented on project '{project.name}': {content[:50]}...")

        serializer = ProjectCommentSerializer(comment, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ProjectDocumentAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        docs = project.project_documents.all()
        serializer = ProjectDocumentSerializer(docs, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        file_obj = request.FILES.get('file')
        name = request.data.get('name') or (file_obj.name if file_obj else '')

        if not file_obj:
            return Response({"detail": "File is required."}, status=status.HTTP_400_BAD_REQUEST)

        doc = ProjectDocument.objects.create(
            project=project,
            name=name,
            file=file_obj,
            uploaded_by=request.user
        )

        log_project_action(project, "File Uploaded", request.user, f"Uploaded document: {name}")

        # Notify manager / admins
        recipients = set()
        if project.assigned_manager and project.assigned_manager != request.user:
            recipients.add(project.assigned_manager)
        
        for r in recipients:
            notify_user(r, "Project File Uploaded", f"{request.user.username} uploaded a document '{name}' to project '{project.name}'.")

        serializer = ProjectDocumentSerializer(doc, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ProjectDashboardAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        role = user.role

        # Base project filtering
        if role in ['HR', 'MD']:
            projects = Project.objects.filter(is_archived=False)
        elif role == 'Manager':
            projects = Project.objects.filter(assigned_manager=user, is_archived=False)
        elif role == 'TeamLead':
            projects = Project.objects.filter(Q(assigned_teams__lead=user), is_archived=False)
        elif role == 'Employee':
            projects = Project.objects.filter(Q(assigned_teams__members=user), is_archived=False)
        else:
            projects = Project.objects.none()

        projects = projects.distinct()

        # Counts
        total_projects = projects.count()
        active_projects = projects.filter(status__in=['Active', 'Assigned']).count()
        completed_projects = projects.filter(status='Completed').count()
        pending_projects = projects.filter(status='Pending').count()
        delayed_projects = projects.filter(status='Delayed').count()

        # Teams count in user scope
        if role in ['HR', 'MD']:
            total_teams = Team.objects.filter(is_active=True).count()
            total_employees = User.objects.filter(role='Employee', is_active=True).count()
        else:
            if role == 'TeamLead':
                assigned_team_ids = Team.objects.filter(lead=user).values_list('id', flat=True)
            elif role == 'Employee':
                assigned_team_ids = Team.objects.filter(members=user).values_list('id', flat=True)
            elif role == 'Manager':
                assigned_team_ids = Team.objects.filter(projects__assigned_manager=user).values_list('id', flat=True)
            else:
                assigned_team_ids = []
            
            reverse_team_ids = Team.objects.filter(projects__in=projects).values_list('id', flat=True)
            combined_team_ids = set(list(assigned_team_ids) + list(reverse_team_ids))
            
            total_teams = Team.objects.filter(id__in=combined_team_ids, is_active=True).distinct().count()
            total_employees = User.objects.filter(teams__id__in=combined_team_ids, is_active=True).distinct().count()

        # Recent activities (logs in scope)
        if role in ['HR', 'MD']:
            logs = ProjectAuditLog.objects.all()[:10]
        else:
            logs = ProjectAuditLog.objects.filter(project__in=projects)[:10]
        
        recent_activities = ProjectAuditLogSerializer(logs, many=True).data

        # Project deadlines in 30 days
        soon = timezone.localdate() + timezone.timedelta(days=30)
        upcoming = projects.filter(deadline__gte=timezone.localdate(), deadline__lte=soon).order_by('deadline')[:5]
        upcoming_deadlines = ProjectSerializer(upcoming, many=True).data

        # Project progress lists (for charts/analytics)
        progress_data = []
        for p in projects:
            # Calculate progress based on Tasks
            tasks_total = Task.objects.filter(project=p).count()
            tasks_completed = Task.objects.filter(project=p, status='Completed').count()
            progress_pct = round((tasks_completed / tasks_total * 100), 2) if tasks_total > 0 else 0.00
            
            progress_data.append({
                "id": p.id,
                "project_code": p.project_code or p.project_id,
                "name": p.name,
                "status": p.status,
                "priority": p.priority,
                "deadline": p.deadline,
                "color": p.project_color,
                "progress": progress_pct
            })

        # Team performance
        team_performance = []
        if role in ['HR', 'MD']:
            teams_to_query = Team.objects.filter(is_active=True)
        else:
            if role == 'TeamLead':
                assigned_team_ids = Team.objects.filter(lead=user).values_list('id', flat=True)
            elif role == 'Employee':
                assigned_team_ids = Team.objects.filter(members=user).values_list('id', flat=True)
            elif role == 'Manager':
                assigned_team_ids = Team.objects.filter(projects__assigned_manager=user).values_list('id', flat=True)
            else:
                assigned_team_ids = []
                
            reverse_team_ids = Team.objects.filter(projects__in=projects).values_list('id', flat=True)
            combined_team_ids = set(list(assigned_team_ids) + list(reverse_team_ids))
            teams_to_query = Team.objects.filter(id__in=combined_team_ids, is_active=True)
        teams_to_query = teams_to_query.distinct()[:5]

        for t in teams_to_query:
            # Get the project associated with this team (check both fields)
            team_project = t.projects.first()
            
            # Tasks completed by team member in team projects
            if team_project:
                team_tasks = Task.objects.filter(project=team_project, assigned_to__in=t.members.all()).distinct()
            else:
                team_tasks = Task.objects.none()
                
            total_t_tasks = team_tasks.count()
            comp_t_tasks = team_tasks.filter(status='Completed').count()
            perf = round((comp_t_tasks / total_t_tasks * 100), 1) if total_t_tasks > 0 else 0
            
            team_performance.append({
                "team_name": t.name,
                "project_name": team_project.name if team_project else "No Project",
                "lead_name": f"{t.lead.first_name} {t.lead.last_name}".strip() if t.lead else "None",
                "members_count": t.members.count(),
                "performance_pct": perf
            })

        data = {
            "total_projects": total_projects,
            "active_projects": active_projects,
            "completed_projects": completed_projects,
            "pending_projects": pending_projects,
            "delayed_projects": delayed_projects,
            "total_teams": total_teams,
            "total_employees": total_employees,
            "recent_activities": recent_activities,
            "upcoming_deadlines": upcoming_deadlines,
            "project_progress": progress_data,
            "team_performance": team_performance
        }

        return Response(data, status=status.HTTP_200_OK)


class ProjectReportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        role = user.role

        # Only MD, HR, or Manager can access reports
        if role not in ['HR', 'MD', 'Manager']:
            return Response({"detail": "Access Denied: You do not have permissions to view reports."}, status=status.HTTP_403_FORBIDDEN)

        # Scope projects
        if role in ['HR', 'MD']:
            projects = Project.objects.all()
        else:
            projects = Project.objects.filter(assigned_manager=user)

        projects = projects.distinct()

        # Build reports JSON structure
        project_summaries = []
        for p in projects:
            tasks_total = Task.objects.filter(project=p).count()
            tasks_completed = Task.objects.filter(project=p, status='Completed').count()
            progress_pct = round((tasks_completed / tasks_total * 100), 2) if tasks_total > 0 else 0.00

            project_summaries.append({
                "project_code": p.project_code or p.project_id,
                "name": p.name,
                "client": p.client_name or (p.client.name if p.client else "Internal"),
                "manager": f"{p.assigned_manager.first_name} {p.assigned_manager.last_name}".strip() if p.assigned_manager else "Unassigned",
                "teams_count": p.assigned_teams.count(),
                "status": p.status,
                "priority": p.priority,
                "budget": str(p.estimated_budget) if p.estimated_budget else "0.00",
                "progress": progress_pct,
                "deadline": str(p.deadline) if p.deadline else "N/A"
            })

        team_summaries = []
        teams = Team.objects.filter(projects__in=projects).distinct()
        for t in teams:
            team_summaries.append({
                "team_code": t.team_code or str(t.id),
                "name": t.name,
                "project_name": t.project.name if t.project else "N/A",
                "leader": f"{t.lead.first_name} {t.lead.last_name}".strip() if t.lead else "None",
                "department": t.get_department_display(),
                "max_size": t.max_size,
                "members_count": t.members_count() if hasattr(t, 'members_count') else t.members.count()
            })

        allocations = []
        # Distinct employees
        employees = User.objects.filter(role='Employee', is_active=True)
        for emp in employees:
            assigned_team = Team.objects.filter(members=emp).first()
            allocated = "Allocated" if assigned_team else "Free"
            allocations.append({
                "emp_id": emp.emp_id,
                "name": f"{emp.first_name} {emp.last_name}".strip() or emp.username,
                "department": emp.get_department_display(),
                "designation": emp.designation or "Employee",
                "status": allocated,
                "project": assigned_team.projects.first().name if assigned_team and assigned_team.projects.first() else "N/A",
                "team": assigned_team.name if assigned_team else "N/A",
                "manager": f"{emp.reporting_manager.first_name} {emp.reporting_manager.last_name}".strip() if emp.reporting_manager else "N/A",
                "team_lead": f"{assigned_team.lead.first_name} {assigned_team.lead.last_name}".strip() if assigned_team and assigned_team.lead else "N/A"
            })

        # Check export parameter
        export_format = request.query_params.get('export', '')

        if export_format == 'excel':
            return self.export_to_excel(project_summaries, team_summaries, allocations)
        elif export_format == 'pdf':
            return self.export_to_pdf(project_summaries, team_summaries, allocations)

        return Response({
            "project_summaries": project_summaries,
            "team_summaries": team_summaries,
            "allocations": allocations
        }, status=status.HTTP_200_OK)

    def export_to_excel(self, project_summaries, team_summaries, allocations):
        wb = Workbook()

        # Sheet 1: Projects Summary
        ws1 = wb.active
        ws1.title = "Project Summary"
        headers1 = ["Project Code", "Project Name", "Client", "Project Manager", "Teams Count", "Status", "Priority", "Budget", "Progress (%)", "Deadline"]
        ws1.append(headers1)
        for p in project_summaries:
            ws1.append([p["project_code"], p["name"], p["client"], p["manager"], p["teams_count"], p["status"], p["priority"], p["budget"], p["progress"], p["deadline"]])

        # Sheet 2: Teams Summary
        ws2 = wb.create_sheet(title="Team Summary")
        headers2 = ["Team Code", "Team Name", "Project", "Team Leader", "Department", "Max Size", "Current Members"]
        ws2.append(headers2)
        for t in team_summaries:
            ws2.append([t["team_code"], t["name"], t["project_name"], t["leader"], t["department"], t["max_size"], t["members_count"]])

        # Sheet 3: Employee Allocation
        ws3 = wb.create_sheet(title="Employee Allocation")
        headers3 = ["Employee ID", "Name", "Department", "Designation", "Status", "Current Project", "Current Team", "Reporting Manager", "Team Lead"]
        ws3.append(headers3)
        for a in allocations:
            ws3.append([a["emp_id"], a["name"], a["department"], a["designation"], a["status"], a["project"], a["team"], a["manager"], a["team_lead"]])

        # Styling
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="092A49", end_color="092A49", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for ws in [ws1, ws2, ws3]:
            # Style header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
            
            # Autofit column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="Enterprise_Project_Management_Report.xlsx"'
        return response

    def export_to_pdf(self, project_summaries, team_summaries, allocations):
        context = {
            "projects": project_summaries,
            "teams": team_summaries,
            "allocations": allocations,
            "generated_date": timezone.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        html_template = """
        <html>
        <head>
            <style>
                @page {
                    size: A4 landscape;
                    margin: 1cm;
                }
                body {
                    font-family: 'Helvetica', 'Arial', sans-serif;
                    color: #1e293b;
                    font-size: 8pt;
                }
                h1 {
                    font-size: 18pt;
                    color: #092a49;
                    text-transform: uppercase;
                    margin-bottom: 5px;
                }
                .meta {
                    margin-bottom: 20px;
                    font-style: italic;
                    color: #64748b;
                }
                h2 {
                    font-size: 12pt;
                    color: #092a49;
                    border-bottom: 1px solid #cbd5e1;
                    padding-bottom: 4px;
                    margin-top: 15px;
                    margin-bottom: 8px;
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                    margin-bottom: 15px;
                }
                th {
                    background-color: #092a49;
                    color: #ffffff;
                    font-weight: bold;
                    text-align: left;
                    padding: 5px;
                }
                td {
                    border-bottom: 1px solid #e2e8f0;
                    padding: 5px;
                }
                .progress-bar {
                    background-color: #e2e8f0;
                    border-radius: 4px;
                    width: 60px;
                    height: 8px;
                    display: inline-block;
                }
                .progress-fill {
                    background-color: #3b82f6;
                    height: 8px;
                    border-radius: 4px;
                }
                .badge {
                    padding: 2px 6px;
                    border-radius: 10px;
                    font-weight: bold;
                    font-size: 7pt;
                }
                .badge-active { background-color: #d1fae5; color: #065f46; }
                .badge-pending { background-color: #fef3c7; color: #92400e; }
                .badge-completed { background-color: #dbeafe; color: #1e40af; }
                .badge-delayed { background-color: #fee2e2; color: #991b1b; }
            </style>
        </head>
        <body>
            <h1>Enterprise Project Management Module Report</h1>
            <div class="meta">Generated on: {{ generated_date }}</div>

            <h2>Project Summary</h2>
            <table>
                <thead>
                    <tr>
                        <th>Code</th>
                        <th>Name</th>
                        <th>Client</th>
                        <th>Manager</th>
                        <th>Teams</th>
                        <th>Status</th>
                        <th>Priority</th>
                        <th>Budget ($)</th>
                        <th>Progress</th>
                        <th>Deadline</th>
                    </tr>
                </thead>
                <tbody>
                    {% for p in projects %}
                    <tr>
                        <td>{{ p.project_code }}</td>
                        <td><b>{{ p.name }}</b></td>
                        <td>{{ p.client }}</td>
                        <td>{{ p.manager }}</td>
                        <td>{{ p.teams_count }}</td>
                        <td><span class="badge badge-{{ p.status|lower }}">{{ p.status }}</span></td>
                        <td>{{ p.priority }}</td>
                        <td>{{ p.budget }}</td>
                        <td>{{ p.progress }}%</td>
                        <td>{{ p.deadline }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>

            <div style="page-break-before: always;"></div>

            <h2>Team Summary</h2>
            <table>
                <thead>
                    <tr>
                        <th>Code</th>
                        <th>Team Name</th>
                        <th>Project</th>
                        <th>Team Leader</th>
                        <th>Department</th>
                        <th>Max Size</th>
                        <th>Current Members</th>
                    </tr>
                </thead>
                <tbody>
                    {% for t in teams %}
                    <tr>
                        <td>{{ t.team_code }}</td>
                        <td><b>{{ t.name }}</b></td>
                        <td>{{ t.project_name }}</td>
                        <td>{{ t.leader }}</td>
                        <td>{{ t.department }}</td>
                        <td>{{ t.max_size }}</td>
                        <td>{{ t.members_count }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>

            <h2>Employee Allocation Report</h2>
            <table>
                <thead>
                    <tr>
                        <th>Employee ID</th>
                        <th>Name</th>
                        <th>Department</th>
                        <th>Designation</th>
                        <th>Allocation Status</th>
                        <th>Current Project</th>
                        <th>Current Team</th>
                        <th>Reporting Manager</th>
                    </tr>
                </thead>
                <tbody>
                    {% for a in allocations %}
                    <tr>
                        <td>{{ a.emp_id }}</td>
                        <td><b>{{ a.name }}</b></td>
                        <td>{{ a.department }}</td>
                        <td>{{ a.designation }}</td>
                        <td>{{ a.status }}</td>
                        <td>{{ a.project }}</td>
                        <td>{{ a.team }}</td>
                        <td>{{ a.manager }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </body>
        </html>
        """

        # Render HTML string
        # Instead of rendering from template file directly, render from string
        from django.template import Template, Context
        template = Template(html_template)
        html = template.render(Context(context))

        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html, dest=pdf_buffer, encoding='UTF-8')

        if pisa_status.err:
            return Response({"detail": "Error generating PDF report."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        pdf_buffer.seek(0)
        response = FileResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Enterprise_Project_Management_Report.pdf"'
        return response


class NotificationAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        notifications = Notification.objects.filter(recipient=user).filter(
            Q(title__in=[
                "Project Created", "Manager Assigned", "Project Deleted",
                "Project Archived Toggle", "Project Transferred", "Project Transferred Away",
                "Team Created", "Team Leader Assigned", "Team Leader Changed",
                "Project Assigned", "Project Removed", "New Project Comment",
                "Project File Uploaded", "Project Updated", "Project Completed"
            ]) | Q(title__icontains="project") | Q(title__icontains="team") | Q(title__icontains="task") | Q(title__icontains="manager")
        ).order_by('-created_at')[:20]
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        user = request.user
        mark_all = request.data.get('mark_all', False)
        notif_id = request.data.get('id')

        if mark_all:
            Notification.objects.filter(recipient=user, is_read=False).update(is_read=True)
            return Response({"detail": "All notifications marked as read."}, status=status.HTTP_200_OK)
        
        if notif_id:
            notif = get_object_or_404(Notification, id=notif_id, recipient=user)
            notif.is_read = True
            notif.save()
            return Response(NotificationSerializer(notif).data, status=status.HTTP_200_OK)

        return Response({"detail": "id or mark_all parameter is required."}, status=status.HTTP_400_BAD_REQUEST)


class StandaloneTeamListCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.role in ['MD', 'HR', 'Manager']:
            teams = Team.objects.all()
        elif user.role == 'TeamLead':
            teams = Team.objects.filter(lead=user)
        else:
            teams = Team.objects.filter(members=user)
        serializer = TeamSerializer(teams.distinct(), many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        user = request.user
        if user.role not in ['MD', 'HR', 'Manager', 'TeamLead']:
            return Response({"detail": "Access Denied."}, status=status.HTTP_403_FORBIDDEN)

        name = request.data.get('name')
        lead_id = request.data.get('lead')
        department = request.data.get('department', 'python_dev')
        description = request.data.get('description', '')
        max_size = int(request.data.get('max_size', 10))
        members_ids = request.data.getlist('members') if hasattr(request.data, 'getlist') else request.data.get('members', [])

        if not name:
            return Response({"detail": "Team Name is required."}, status=status.HTTP_400_BAD_REQUEST)

        lead = None
        if lead_id:
            try:
                lead = User.objects.get(id=lead_id, role='TeamLead')
            except User.DoesNotExist:
                return Response({"detail": "Selected Team Lead is invalid."}, status=status.HTTP_400_BAD_REQUEST)

        # TeamLead can only create a team for themselves
        if user.role == 'TeamLead':
            lead = user

        team = Team.objects.create(
            name=name,
            lead=lead,
            department=department,
            description=description,
            max_size=max_size
        )
        
        # Add members
        if members_ids:
            if isinstance(members_ids, str):
                members_ids = [members_ids]
            
            for m_id in members_ids:
                try:
                    emp = User.objects.get(id=int(m_id), role='Employee')
                    if team.members.count() < team.max_size:
                        team.members.add(emp)
                        emp.team_name = team.name
                        emp.save()
                except (User.DoesNotExist, ValueError):
                    pass

        notify_user(user, "Team Created", f"Team '{name}' has been successfully created.")
        if lead and lead != user:
            notify_user(lead, "Team Leader Assigned", f"You have been assigned as Team Lead for '{name}'.")

        serializer = TeamSerializer(team)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
class ProjectReviewAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        project = get_object_or_404(Project, id=project_id)
        from .models import ProjectReview
        from .serializers import ProjectReviewSerializer # Need to create this
        review = getattr(project, 'review', None)
        if not review:
            return Response({"detail": "No project review found."}, status=status.HTTP_404_NOT_FOUND)
        return Response(ProjectReviewSerializer(review).data, status=status.HTTP_200_OK)

    def post(self, request, project_id):
        user = request.user
        project = get_object_or_404(Project, id=project_id)
        action = request.data.get('action')
        
        if not action:
            return Response({"detail": "Action is required."}, status=status.HTTP_400_BAD_REQUEST)

        from .models import ProjectReview
        review, created = ProjectReview.objects.get_or_create(project=project)
        
        if user.role == 'TeamLead':
            if not project.assigned_teams.filter(lead=user).exists():
                return Response({"detail": "Only the assigned TL can submit the project report."}, status=status.HTTP_403_FORBIDDEN)
            if review.review_status != 'Pending TL Submission':
                return Response({"detail": "Project is not pending TL submission."}, status=status.HTTP_400_BAD_REQUEST)
                
            review.tl_remarks = request.data.get('remarks', '')
            review.tl_reviewed_by = user
            review.tl_reviewed_at = timezone.now()
            review.review_status = 'Pending Manager'
            review.save()
            
            if project.assigned_manager:
                notify_user(project.assigned_manager, "Project Report Submitted", f"TL {user.username} has submitted the progress report for project '{project.name}'. Pending your review.")
            return Response({"detail": "Project report submitted to Manager."}, status=status.HTTP_200_OK)
            
        elif user.role == 'Manager':
            if project.assigned_manager != user:
                return Response({"detail": "Only the assigned Manager can review."}, status=status.HTTP_403_FORBIDDEN)
            if review.review_status != 'Pending Manager':
                return Response({"detail": "Project is not pending Manager review."}, status=status.HTTP_400_BAD_REQUEST)
                
            review.manager_remarks = request.data.get('remarks', '')
            review.manager_reviewed_by = user
            review.manager_reviewed_at = timezone.now()
            review.review_status = 'Pending HR'
            review.save()
            
            for hr in User.objects.filter(role='HR'):
                notify_user(hr, "Project Progress Pending HR Review", f"Manager {user.username} submitted the progress for project '{project.name}'.")
            return Response({"detail": "Project report submitted to HR."}, status=status.HTTP_200_OK)
            
        elif user.role == 'HR':
            if review.review_status != 'Pending HR':
                return Response({"detail": "Project is not pending HR review."}, status=status.HTTP_400_BAD_REQUEST)
                
            review.hr_remarks = request.data.get('remarks', '')
            review.hr_reviewed_by = user
            review.hr_reviewed_at = timezone.now()
            review.review_status = 'Pending MD'
            review.save()
            
            for md in User.objects.filter(role='MD'):
                notify_user(md, "Project Progress Pending MD Review", f"HR {user.username} submitted the progress for project '{project.name}'.")
            return Response({"detail": "Project report submitted to MD."}, status=status.HTTP_200_OK)
            
        elif user.role == 'MD':
            if review.review_status != 'Pending MD':
                return Response({"detail": "Project is not pending MD review."}, status=status.HTTP_400_BAD_REQUEST)
                
            review.md_remarks = request.data.get('remarks', '')
            review.md_reviewed_by = user
            review.md_reviewed_at = timezone.now()
            review.review_status = 'Finalized'
            review.save()
            
            project.status = 'Completed'
            project.save()
            
            return Response({"detail": "Project completed successfully."}, status=status.HTTP_200_OK)
        else:
            return Response({"detail": "Role not authorized."}, status=status.HTTP_403_FORBIDDEN)
